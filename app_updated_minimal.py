
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime
import os
import io

# ─── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = "Book1.xlsx"
SHEET_NAME = "Sheet2"
COLUMNS = [
    "BugID", "Title", "Description", "AppName",
    "Validation Result", "OS Build", "Current Status",
    "Enigineer", "Date of assign", "Date of complete", "Empty",
    "Category", "Cate Challenges", "Challenges",
    "Remarks", "Used Cate(Yes/No)"
]

CATEGORY_OPTIONS = ["Anomaly Detection", "aetriage", "Compatemerging", "OCVRegularSerach","Other"]
USED_CATE_OPTIONS = ["", "Yes", "No"]

VALIDATION_OPTIONS = ["Repro", "Not Repro"]
STATUS_OPTIONS = ["Completed", "On Hold", "Pending", "Yet to Start", "In Progress"]
ENGINEERS = ["Thomas", "Anshika", "John"]

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bug Tracker",
    page_icon="🐛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Excel Helpers ────────────────────────────────────────────────────────────

def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)

def read_data() -> pd.DataFrame:
    ensure_excel_exists()
    df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, dtype={"BugID": "Int64"})
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNS]

def append_row(row_data: dict):
    ensure_excel_exists()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([row_data.get(col, None) for col in COLUMNS])
    wb.save(EXCEL_FILE)

def update_row(bug_id: int, row_data: dict):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(bug_id):
            for col_name, value in row_data.items():
                if col_name in header:
                    row[header.index(col_name)].value = value
    wb.save(EXCEL_FILE)

def delete_row(bug_id: int):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(bug_id):
            ws.delete_rows(i)
            break
    wb.save(EXCEL_FILE)

def get_excel_bytes():
    with open(EXCEL_FILE, "rb") as f:
        return f.read()

# ─── UI ───────────────────────────────────────────────────────────────────────
with st.sidebar:
    nav = st.radio("Navigation", ["📊 Dashboard", "➕ Add Bug", "✏️ Edit Bug", "🗑️ Delete Bug"])

df = read_data()

# ─── ADD BUG ──────────────────────────────────────────────────────────────────
if nav == "➕ Add Bug":
    st.title("Add Bug")

    with st.form("add_form"):
        c1, c2 = st.columns(2)

        with c1:
            bug_id = st.text_input("BugID")
            title = st.text_input("Title")
            app_name = st.text_input("AppName")
            engineer = st.selectbox("Enigineer", ENGINEERS)
            category = st.selectbox("Category", CATEGORY_OPTIONS)
            date_assign = None

        with c2:
            val_result = st.selectbox("Validation Result", VALIDATION_OPTIONS)
            os_build = st.text_input("OS Build")
            current_status = st.selectbox("Current Status", STATUS_OPTIONS)
            description = st.text_area("Description")
            challenges = st.text_area("Challenges")
            remarks = st.text_input("Remarks")
            used_cate = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS)
            date_complete = st.date_input("Date of complete", value=None)

        cate_challenges = ""
        if used_cate == "No":
            cate_challenges = st.text_area("Cate Challenges")

        submitted = st.form_submit_button("Add")

        if submitted:
            row = {
                "BugID": int(bug_id),
                "Title": title,
                "Description": description,
                "AppName": app_name,
                "Validation Result": val_result,
                "OS Build": os_build,
                "Current Status": current_status,
                "Enigineer": engineer,
                "Date of assign": None,
                "Date of complete": date_complete.strftime("%Y-%m-%d") if date_complete else None,
                "Empty": None,
                "Category": category,
                "Cate Challenges": cate_challenges if used_cate == "No" else None,
                "Challenges": challenges,
                "Remarks": remarks,
                "Used Cate(Yes/No)": used_cate,
            }
            append_row(row)
            st.success("Bug Added")

# ─── EDIT BUG ─────────────────────────────────────────────────────────────────
elif nav == "✏️ Edit Bug":
    st.title("Edit Bug")

    bug_ids = df["BugID"].dropna().astype(int).tolist()
    selected_id = st.selectbox("BugID", bug_ids)

    row = df[df["BugID"] == selected_id].iloc[0]

    with st.form("edit_form"):
        val_result = st.selectbox("Validation Result", VALIDATION_OPTIONS)
        os_build = st.text_input("OS Build", row.get("OS Build",""))
        current_status = st.selectbox("Current Status", STATUS_OPTIONS)
        engineer = st.selectbox("Enigineer", ENGINEERS)
        used_cate = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS)

        cate_challenges = ""
        if used_cate == "No":
            cate_challenges = st.text_area("Cate Challenges", value=row.get("Cate Challenges",""))

        date_assign = st.date_input("Date of assign")

        save = st.form_submit_button("Save")

        if save:
            update_row(selected_id,{
                "Validation Result": val_result,
                "OS Build": os_build,
                "Current Status": current_status,
                "Enigineer": engineer,
                "Date of assign": date_assign.strftime("%Y-%m-%d"),
                "Cate Challenges": cate_challenges,
                "Used Cate(Yes/No)": used_cate
            })
            st.success("Updated")

# ─── DASHBOARD ────────────────────────────────────────────────────────────────
elif nav == "📊 Dashboard":
    st.dataframe(df)

# ─── DELETE ───────────────────────────────────────────────────────────────────
elif nav == "🗑️ Delete Bug":
    bug_ids = df["BugID"].dropna().astype(int).tolist()
    del_id = st.selectbox("BugID", bug_ids)

    if st.button("Delete"):
        delete_row(del_id)
        st.success("Deleted")
