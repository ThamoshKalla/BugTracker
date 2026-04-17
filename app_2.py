import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime
import os
import io

# ─── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = "Book1.xlsx"
SHEET_NAME = "Sheet2"
COLUMNS = [
    "BugID", "Title", "Description", "AppName", "Validation Result",
    "Enigineer", "Date of assign", "Date of complete", "Empty",
    "Category", "Challenges", "Remarks", "Used Cate(Yes/No)",
    "OSBuild", "Current Status", "Cate Challenges"
]
CATEGORY_OPTIONS   = ["", "Anomaly Detection", "aetriage", "Compatemerging", "OCVRegularSearch", "Other"]
USED_CATE_OPTIONS  = ["", "Yes", "No"]
ENGINEER_OPTIONS   = ["", "Thomas", "Anshika", "Sai Ram", "Lahari", "Kiran"]
VALIDATION_OPTIONS = ["", "Repro - Regression", "NotRepro", "Repro - Non-Regression" , "N/A"]
STATUS_OPTIONS     = ["", "Yet to Start", "In Progress", "Pending", "On Hold", "Completed"]

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bug Tracker",
    page_icon="🐛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Syne:wght@400;600;700;800&display=swap');

:root {
    --bg: #0d0f14;
    --surface: #161a23;
    --surface2: #1e2433;
    --border: #2a3144;
    --accent: #e8ff47;
    --accent2: #ff4757;
    --accent3: #00d4ff;
    --text: #e8eaf0;
    --muted: #6b7280;
    --pending: #ff4757;
    --completed: #00c896;
}

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: var(--bg);
    color: var(--text);
}

[data-testid="stSidebar"] {
    background-color: var(--surface) !important;
    border-right: 1px solid var(--border);
}
[data-testid="stSidebar"] .stRadio label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: var(--text) !important;
}

.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 16px;
    margin-bottom: 28px;
}
.kpi-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px 24px;
    position: relative;
    overflow: hidden;
}
.kpi-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.kpi-card.total::before     { background: var(--accent3); }
.kpi-card.completed::before { background: var(--completed); }
.kpi-card.pending::before   { background: var(--pending); }
.kpi-card.category::before  { background: var(--accent); }

.kpi-label {
    font-size: 11px; font-weight: 700; letter-spacing: 2px;
    text-transform: uppercase; color: var(--muted); margin-bottom: 8px;
    font-family: 'JetBrains Mono', monospace;
}
.kpi-value {
    font-size: 38px; font-weight: 800; line-height: 1;
    font-family: 'Syne', sans-serif;
}
.kpi-card.total .kpi-value     { color: var(--accent3); }
.kpi-card.completed .kpi-value { color: var(--completed); }
.kpi-card.pending .kpi-value   { color: var(--pending); }
.kpi-card.category .kpi-value  { color: var(--accent); }
.kpi-sub { font-size: 12px; color: var(--muted); margin-top: 4px; font-family: 'JetBrains Mono', monospace; }

.section-header {
    font-size: 22px; font-weight: 800; letter-spacing: -0.5px;
    margin: 28px 0 16px 0; display: flex; align-items: center; gap: 10px;
    border-bottom: 1px solid var(--border); padding-bottom: 10px;
}
.section-header .dot {
    width: 8px; height: 8px; border-radius: 50%;
    background: var(--accent); display: inline-block;
}

.stTextInput input, .stNumberInput input, .stSelectbox select,
.stTextArea textarea, .stDateInput input {
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 13px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(232,255,71,0.15) !important;
}
label[data-testid="stWidgetLabel"] {
    color: var(--muted) !important; font-size: 11px !important;
    font-weight: 700 !important; letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    font-family: 'JetBrains Mono', monospace !important;
}

.stButton button {
    background: var(--accent) !important; color: #0d0f14 !important;
    font-weight: 700 !important; font-family: 'Syne', sans-serif !important;
    border: none !important; border-radius: 8px !important;
    font-size: 13px !important; letter-spacing: 0.5px; transition: all 0.2s;
}
.stButton button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 20px rgba(232,255,71,0.3) !important;
}
.stDownloadButton button {
    background: var(--surface) !important; color: var(--accent3) !important;
    border: 1px solid var(--accent3) !important; font-weight: 700 !important;
    font-family: 'Syne', sans-serif !important; border-radius: 8px !important;
}

.stCheckbox label {
    color: var(--text) !important;
    font-family: 'JetBrains Mono', monospace !important; font-size: 13px !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--border); border-radius: 10px; overflow: hidden;
}

.app-title {
    font-size: 36px; font-weight: 800; letter-spacing: -1px; margin-bottom: 4px;
}
.app-subtitle {
    font-size: 14px; color: var(--muted);
    font-family: 'JetBrains Mono', monospace; margin-bottom: 28px;
}
.accent-text { color: var(--accent); }

div[data-baseweb="select"] > div {
    background: var(--surface2) !important;
    border-color: var(--border) !important; color: var(--text) !important;
}
div[data-baseweb="select"] span { color: var(--text) !important; }

hr { border-color: var(--border) !important; }

.streamlit-expanderHeader {
    background: var(--surface) !important; border: 1px solid var(--border) !important;
    border-radius: 8px !important; color: var(--text) !important;
    font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
}

.block-container { padding-top: 2rem !important; }

.badge-pending {
    display: inline-block; background: rgba(255,71,87,0.15); color: var(--pending);
    border: 1px solid rgba(255,71,87,0.4); border-radius: 4px; padding: 2px 8px;
    font-size: 11px; font-weight: 700; font-family: 'JetBrains Mono', monospace; letter-spacing: 1px;
}
.badge-done {
    display: inline-block; background: rgba(0,200,150,0.15); color: var(--completed);
    border: 1px solid rgba(0,200,150,0.4); border-radius: 4px; padding: 2px 8px;
    font-size: 11px; font-weight: 700; font-family: 'JetBrains Mono', monospace; letter-spacing: 1px;
}
.cate-note {
    background: rgba(232,255,71,0.07); border: 1px solid rgba(232,255,71,0.25);
    border-radius: 8px; padding: 10px 14px; font-size: 12px; color: #e8ff47;
    font-family: 'JetBrains Mono', monospace; margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def ensure_excel_exists():
    """Create the Excel file with headers if missing; add new columns if needed."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
        return
    wb = load_workbook(EXCEL_FILE)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
        return
    ws = wb[SHEET_NAME]
    existing = [cell.value for cell in ws[1]]
    added = False
    for col in COLUMNS:
        if col not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=col)
            existing.append(col)
            added = True
    if added:
        wb.save(EXCEL_FILE)


def read_data() -> pd.DataFrame:
    ensure_excel_exists()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, dtype={"BugID": "Int64"})
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = None
        return df[COLUMNS]
    except Exception as e:
        st.error(f"Error reading Excel: {e}")
        return pd.DataFrame(columns=COLUMNS)


def append_row(row_data: dict):
    ensure_excel_exists()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([row_data.get(col, None) for col in COLUMNS])
    wb.save(EXCEL_FILE)


def update_row(bug_id: int, row_data: dict):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        try:
            if int(row[0].value) == bug_id:
                for col_name, value in row_data.items():
                    if col_name in header:
                        row[header.index(col_name)].value = value
                break
        except (TypeError, ValueError):
            continue
    wb.save(EXCEL_FILE)


def delete_row(bug_id: int):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            if int(row[0].value) == bug_id:
                ws.delete_rows(i)
                break
        except (TypeError, ValueError):
            continue
    wb.save(EXCEL_FILE)


def get_excel_bytes() -> bytes:
    with open(EXCEL_FILE, "rb") as f:
        return f.read()


def safe_date(val):
    if pd.isna(val) or val is None or str(val).strip() in ("", "NaT", "None", "nan"):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def safe_idx(options, val, default=0):
    try:
        return options.index(val) if val in options else default
    except Exception:
        return default


# ─── Session State ────────────────────────────────────────────────────────────
if "refresh" not in st.session_state:
    st.session_state.refresh = 0


def trigger_refresh():
    st.session_state.refresh += 1


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:16px 0 24px 0;">
        <div style="font-size:24px;font-weight:800;letter-spacing:-0.5px;">🐛 BugTracker</div>
        <div style="font-size:11px;color:#6b7280;font-family:'JetBrains Mono',monospace;margin-top:4px;">v2.0 · openpyxl engine</div>
    </div>
    """, unsafe_allow_html=True)

    nav = st.radio(
        "Navigation",
        ["📊 Dashboard", "➕ Add Bug", "✏️ Edit Bug", "🗑️ Delete Bug"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:11px;color:#6b7280;font-family:'JetBrains Mono',monospace;line-height:1.8;">
    FILE<br><span style="color:#e8eaf0;">{EXCEL_FILE}</span><br><br>
    SHEET<br><span style="color:#e8eaf0;">{SHEET_NAME}</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    ensure_excel_exists()
    st.download_button(
        "⬇️ Download Excel",
        data=get_excel_bytes(),
        file_name="Book1.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


# ─── Load Data ────────────────────────────────────────────────────────────────
df = read_data()
total     = len(df)
completed = int(df["Date of complete"].notna().sum()) if total > 0 else 0
pending   = total - completed
cat_counts = df["Category"].value_counts().to_dict() if total > 0 else {}


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if nav == "📊 Dashboard":
    st.markdown("""
    <div class="app-title">Bug <span class="accent-text">Tracker</span></div>
    <div class="app-subtitle">Real-time · Excel-backed · openpyxl engine</div>
    """, unsafe_allow_html=True)

    cat_summary_text = " · ".join([f"{k}: {v}" for k, v in cat_counts.items()]) or "—"
    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card total">
            <div class="kpi-label">Total Bugs</div>
            <div class="kpi-value">{total}</div>
            <div class="kpi-sub">all records</div>
        </div>
        <div class="kpi-card completed">
            <div class="kpi-label">Completed</div>
            <div class="kpi-value">{completed}</div>
            <div class="kpi-sub">{round(completed/total*100) if total else 0}% done</div>
        </div>
        <div class="kpi-card pending">
            <div class="kpi-label">Pending</div>
            <div class="kpi-value">{pending}</div>
            <div class="kpi-sub">no completion date</div>
        </div>
        <div class="kpi-card category">
            <div class="kpi-label">Categories</div>
            <div class="kpi-value">{len(cat_counts)}</div>
            <div class="kpi-sub">{cat_summary_text}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-header"><span class="dot"></span> Records</div>', unsafe_allow_html=True)
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        f_bugid = st.text_input("🔍 BugID", placeholder="e.g. 61354949")
    with fc2:
        f_app = st.text_input("🔍 AppName", placeholder="filter by app")
    with fc3:
        f_eng = st.text_input("🔍 Engineer", placeholder="filter by name")
    with fc4:
        f_cat = st.selectbox("🔍 Category", ["All"] + [c for c in CATEGORY_OPTIONS if c])

    filtered = df.copy()
    if f_bugid.strip():
        try:
            filtered = filtered[filtered["BugID"] == int(f_bugid.strip())]
        except ValueError:
            pass
    if f_app.strip():
        filtered = filtered[filtered["AppName"].astype(str).str.contains(f_app.strip(), case=False, na=False)]
    if f_eng.strip():
        filtered = filtered[filtered["Enigineer"].astype(str).str.contains(f_eng.strip(), case=False, na=False)]
    if f_cat != "All":
        filtered = filtered[filtered["Category"] == f_cat]

    def style_rows(row):
        is_pending = pd.isna(row["Date of complete"]) or str(row["Date of complete"]).strip() in ("", "NaT", "None", "nan")
        if is_pending:
            return ["background-color: rgba(255,71,87,0.08); color: #ffd0d4"] * len(row)
        return ["background-color: rgba(0,200,150,0.05); color: #e8eaf0"] * len(row)

    if not filtered.empty:
        st.dataframe(filtered.style.apply(style_rows, axis=1), use_container_width=True, height=420)
        st.markdown(f"""
        <div style="font-size:12px;color:#6b7280;font-family:'JetBrains Mono',monospace;margin-top:8px;">
        Showing {len(filtered)} of {total} records &nbsp;·&nbsp;
        <span style="color:#ff4757;">■</span> Pending &nbsp;
        <span style="color:#00c896;">■</span> Completed
        </div>""", unsafe_allow_html=True)
    else:
        st.info("No records found matching your filters.")


# ══════════════════════════════════════════════════════════════════════════════
# ADD BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "➕ Add Bug":
    st.markdown("""
    <div class="app-title">Add <span class="accent-text">New Bug</span></div>
    <div class="app-subtitle">Fill in the form to append a record to the Excel file</div>
    """, unsafe_allow_html=True)

    with st.form("add_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)

        with c1:
            bug_id   = st.text_input("BugID *", placeholder="e.g. 61354949")
            title    = st.text_input("Title")
            app_name = st.text_input("AppName")
            engineer = st.selectbox("Enigineer", ENGINEER_OPTIONS)

        with c2:
            val_result     = st.selectbox("Validation Result", VALIDATION_OPTIONS)
            category       = st.selectbox("Category", CATEGORY_OPTIONS)
            used_cate      = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS)
            current_status = st.selectbox("Current Status", STATUS_OPTIONS)

        with c3:
            osbuild     = st.text_input("OSBuild", placeholder="e.g. 26100.1234")
            remarks     = st.text_input("Remarks")
            #empty_field = st.text_input("Empty (optional)")
            # Date of assign: blank by default
            date_assign = st.date_input("Date of assign (optional)", value=None)

        ta1, ta2 = st.columns(2)
        with ta1:
            description = st.text_area("Description", height=90)
            challenges  = st.text_area("Challenges", height=68)
        with ta2:
            st.markdown(
                '<div class="cate-note">⚠️ Fill Cate Challenges when Used Cate = No</div>'
                if used_cate == "No" else "",
                unsafe_allow_html=True
            )
            cate_challenges = st.text_area("Cate Challenges", height=90,
                                           help="Required when Used Cate(Yes/No) = No")
            date_complete = st.date_input("Date of complete (optional)", value=None)

        submitted = st.form_submit_button("➕ Add Record", use_container_width=True)

    if submitted:
        if not bug_id.strip():
            st.toast("⚠️ BugID is required.", icon="⚠️")
        else:
            try:
                bid_int = int(bug_id.strip())
            except ValueError:
                st.toast("⚠️ BugID must be a number.", icon="⚠️")
                st.stop()

            if bid_int in df["BugID"].dropna().astype(int).tolist():
                st.toast(f"⚠️ BugID {bid_int} already exists.", icon="⚠️")
            else:
                new_row = {
                    "BugID": bid_int,
                    "Title": title,
                    "Description": description,
                    "AppName": app_name,
                    "Validation Result": val_result or None,
                    "Enigineer": engineer or None,
                    "Date of assign": date_assign.strftime("%Y-%m-%d") if date_assign else None,
                    "Date of complete": date_complete.strftime("%Y-%m-%d") if date_complete else None,
                   # "Empty": empty_field or None,
                    "Category": category or None,
                    "Challenges": challenges or None,
                    "Remarks": remarks or None,
                    "Used Cate(Yes/No)": used_cate or None,
                    "OSBuild": osbuild or None,
                    "Current Status": current_status or None,
                    "Cate Challenges": cate_challenges or None,
                }
                try:
                    append_row(new_row)
                    st.toast(f"✅ Bug #{bid_int} added successfully!", icon="✅")
                    trigger_refresh()
                    st.rerun()
                except Exception as e:
                    st.toast(f"❌ Error: {e}", icon="❌")


# ══════════════════════════════════════════════════════════════════════════════
# EDIT BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "✏️ Edit Bug":
    st.markdown("""
    <div class="app-title">Edit <span class="accent-text">Bug</span></div>
    <div class="app-subtitle">Select a record by BugID and update its values in-place</div>
    """, unsafe_allow_html=True)

    if df.empty:
        st.info("No records available to edit.")
    else:
        bug_ids = sorted(df["BugID"].dropna().astype(int).tolist())
        selected_id = st.selectbox("Select BugID", bug_ids)
        row = df[df["BugID"] == selected_id].iloc[0]

        def sv(col):
            """Safe string value."""
            v = row.get(col)
            return str(v) if pd.notna(v) else ""

        with st.form("edit_form"):
            c1, c2, c3 = st.columns(3)

            with c1:
                title    = st.text_input("Title", value=sv("Title"))
                app_name = st.text_input("AppName", value=sv("AppName"))
                engineer = st.selectbox("Enigineer", ENGINEER_OPTIONS,
                                        index=safe_idx(ENGINEER_OPTIONS, sv("Enigineer")))
                category = st.selectbox("Category", CATEGORY_OPTIONS,
                                        index=safe_idx(CATEGORY_OPTIONS, sv("Category")))

            with c2:
                val_result = st.selectbox("Validation Result", VALIDATION_OPTIONS,
                                          index=safe_idx(VALIDATION_OPTIONS, sv("Validation Result")))
                used_cate  = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS,
                                          index=safe_idx(USED_CATE_OPTIONS, sv("Used Cate(Yes/No)")))
                current_status = st.selectbox("Current Status", STATUS_OPTIONS,
                                              index=safe_idx(STATUS_OPTIONS, sv("Current Status")))
                osbuild = st.text_input("OSBuild", value=sv("OSBuild"))

            with c3:
                remarks     = st.text_input("Remarks", value=sv("Remarks"))
                empty_field = st.text_input("Empty", value=sv("Empty"))
                # Date of assign — editable; blank if not set in record
                date_assign   = st.date_input("Date of assign (optional)",
                                              value=safe_date(row.get("Date of assign")))
                date_complete = st.date_input("Date of complete (optional)",
                                              value=safe_date(row.get("Date of complete")))

            ta1, ta2 = st.columns(2)
            with ta1:
                description = st.text_area("Description", value=sv("Description"), height=90)
                challenges  = st.text_area("Challenges",  value=sv("Challenges"),  height=68)
            with ta2:
                if used_cate == "No":
                    st.markdown('<div class="cate-note">⚠️ Used Cate = No → fill Cate Challenges</div>',
                                unsafe_allow_html=True)
                cate_challenges = st.text_area("Cate Challenges", value=sv("Cate Challenges"), height=90)

            save = st.form_submit_button("💾 Save Changes", use_container_width=True)

        if save:
            updated = {
                "Title": title,
                "Description": description,
                "AppName": app_name,
                "Validation Result": val_result or None,
                "Enigineer": engineer or None,
                "Date of assign": date_assign.strftime("%Y-%m-%d") if date_assign else None,
                "Date of complete": date_complete.strftime("%Y-%m-%d") if date_complete else None,
                "Empty": empty_field or None,
                "Category": category or None,
                "Challenges": challenges or None,
                "Remarks": remarks or None,
                "Used Cate(Yes/No)": used_cate or None,
                "OSBuild": osbuild or None,
                "Current Status": current_status or None,
                "Cate Challenges": cate_challenges or None,
            }
            try:
                update_row(selected_id, updated)
                st.toast(f"✅ Bug #{selected_id} updated!", icon="✅")
                trigger_refresh()
                st.rerun()
            except Exception as e:
                st.toast(f"❌ Error: {e}", icon="❌")


# ══════════════════════════════════════════════════════════════════════════════
# DELETE BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "🗑️ Delete Bug":
    st.markdown("""
    <div class="app-title">Delete <span class="accent-text">Bug</span></div>
    <div class="app-subtitle">Permanently remove a record from the Excel file</div>
    """, unsafe_allow_html=True)

    if df.empty:
        st.info("No records available to delete.")
    else:
        bug_ids = sorted(df["BugID"].dropna().astype(int).tolist())
        del_id  = st.selectbox("Select BugID to delete", bug_ids)

        st.markdown("**Record preview:**")
        st.dataframe(df[df["BugID"] == del_id], use_container_width=True)

        confirm = st.checkbox(f"✅ I confirm I want to permanently delete Bug #{del_id}")

        if st.button("🗑️ Delete Record", disabled=not confirm):
            try:
                delete_row(del_id)
                st.toast(f"🗑️ Bug #{del_id} deleted.", icon="🗑️")
                trigger_refresh()
                st.rerun()
            except Exception as e:
                st.toast(f"❌ Error: {e}", icon="❌")
