import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import date, datetime
import os
import io

# ─── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = "Book1.xlsx"
SHEET_NAME = "Sheet2"
COLUMNS = [
    "BugID", "Title", "Description", "AppName", "Validation Result",
    "Enigineer", "Date of assign", "Date of complete", "Empty",
    "Category", "Challenges", "Remarks", "Used Cate(Yes/No)"
]
CATEGORY_OPTIONS = ["Anomaly Detection", "aetriage", "Compatemerging", "OCVRegularSerach","Other"]
USED_CATE_OPTIONS = ["", "Yes", "No"]

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bug Tracker",
    page_icon="🐛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Syne:wght@400;600;700;800&display=swap');

:root {
    --bg: #0d0f14;
    --surface: #161a23;
    --surface2: #1e2433;
    --border: #2a3144;
    --accent: #e8ff47;
    --accent2: #ff4757;
    --accent3: #00d4ff;
    --text: #e8eaf0;
    --muted: #6b7280;
    --pending: #ff4757;
    --completed: #00c896;
}

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: var(--bg);
    color: var(--text);
}

/* Sidebar */
[data-testid="stSidebar"] {
    background-color: var(--surface) !important;
    border-right: 1px solid var(--border);
}
[data-testid="stSidebar"] .stRadio label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: var(--text) !important;
}

/* KPI Cards */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 16px;
    margin-bottom: 28px;
}
.kpi-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px 24px;
    position: relative;
    overflow: hidden;
}
.kpi-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.kpi-card.total::before  { background: var(--accent3); }
.kpi-card.completed::before { background: var(--completed); }
.kpi-card.pending::before   { background: var(--pending); }
.kpi-card.category::before  { background: var(--accent); }

.kpi-label {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 8px;
    font-family: 'JetBrains Mono', monospace;
}
.kpi-value {
    font-size: 38px;
    font-weight: 800;
    line-height: 1;
    font-family: 'Syne', sans-serif;
}
.kpi-card.total .kpi-value    { color: var(--accent3); }
.kpi-card.completed .kpi-value{ color: var(--completed); }
.kpi-card.pending .kpi-value  { color: var(--pending); }
.kpi-card.category .kpi-value { color: var(--accent); }
.kpi-sub { font-size: 12px; color: var(--muted); margin-top: 4px; font-family: 'JetBrains Mono', monospace;}

/* Section Headers */
.section-header {
    font-size: 22px;
    font-weight: 800;
    letter-spacing: -0.5px;
    margin: 28px 0 16px 0;
    display: flex;
    align-items: center;
    gap: 10px;
    border-bottom: 1px solid var(--border);
    padding-bottom: 10px;
}
.section-header .dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    background: var(--accent);
    display: inline-block;
}

/* Filter bar */
.filter-bar {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 16px;
}

/* Inputs */
.stTextInput input, .stNumberInput input, .stSelectbox select,
.stTextArea textarea, .stDateInput input {
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 13px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(232,255,71,0.15) !important;
}
label[data-testid="stWidgetLabel"] {
    color: var(--muted) !important;
    font-size: 11px !important;
    font-weight: 700 !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    font-family: 'JetBrains Mono', monospace !important;
}

/* Buttons */
.stButton button {
    background: var(--accent) !important;
    color: #0d0f14 !important;
    font-weight: 700 !important;
    font-family: 'Syne', sans-serif !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 13px !important;
    letter-spacing: 0.5px;
    transition: all 0.2s;
}
.stButton button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 20px rgba(232,255,71,0.3) !important;
}
.stDownloadButton button {
    background: var(--surface) !important;
    color: var(--accent3) !important;
    border: 1px solid var(--accent3) !important;
    font-weight: 700 !important;
    font-family: 'Syne', sans-serif !important;
    border-radius: 8px !important;
}

/* Checkbox */
.stCheckbox label {
    color: var(--text) !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 13px !important;
}

/* Dataframe */
[data-testid="stDataFrame"] {
    border: 1px solid var(--border);
    border-radius: 10px;
    overflow: hidden;
}

/* App title */
.app-title {
    font-size: 36px;
    font-weight: 800;
    letter-spacing: -1px;
    margin-bottom: 4px;
}
.app-subtitle {
    font-size: 14px;
    color: var(--muted);
    font-family: 'JetBrains Mono', monospace;
    margin-bottom: 28px;
}
.accent-text { color: var(--accent); }

/* Selectbox */
div[data-baseweb="select"] > div {
    background: var(--surface2) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
}
div[data-baseweb="select"] span {
    color: var(--text) !important;
}

/* Form container */
.form-container {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 24px;
}

/* Divider */
hr { border-color: var(--border) !important; }

/* Expander */
.streamlit-expanderHeader {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text) !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
}

/* Remove default padding */
.block-container { padding-top: 2rem !important; }

/* Pending badge */
.badge-pending {
    display: inline-block;
    background: rgba(255,71,87,0.15);
    color: var(--pending);
    border: 1px solid rgba(255,71,87,0.4);
    border-radius: 4px;
    padding: 2px 8px;
    font-size: 11px;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
    letter-spacing: 1px;
}
.badge-done {
    display: inline-block;
    background: rgba(0,200,150,0.15);
    color: var(--completed);
    border: 1px solid rgba(0,200,150,0.4);
    border-radius: 4px;
    padding: 2px 8px;
    font-size: 11px;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
    letter-spacing: 1px;
}
</style>
""", unsafe_allow_html=True)


# ─── Excel Helpers ────────────────────────────────────────────────────────────

def ensure_excel_exists():
    """Create the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)


def read_data() -> pd.DataFrame:
    """Read data from Excel using pandas."""
    ensure_excel_exists()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, dtype={"BugID": "Int64"})
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df[COLUMNS]
        return df
    except Exception as e:
        st.error(f"Error reading Excel: {e}")
        return pd.DataFrame(columns=COLUMNS)


def append_row(row_data: dict):
    """Append a new row to the Excel file without touching existing formatting."""
    ensure_excel_exists()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    row = [row_data.get(col, None) for col in COLUMNS]
    ws.append(row)
    wb.save(EXCEL_FILE)


def update_row(bug_id: int, row_data: dict):
    """Update an existing row in-place by BugID."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        try:
            if int(row[0].value) == bug_id:
                for col_name, value in row_data.items():
                    if col_name in header:
                        idx = header.index(col_name)
                        row[idx].value = value
                break
        except (TypeError, ValueError):
            continue
    wb.save(EXCEL_FILE)


def delete_row(bug_id: int):
    """Delete a row by BugID."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            if int(row[0].value) == bug_id:
                ws.delete_rows(i)
                break
        except (TypeError, ValueError):
            continue
    wb.save(EXCEL_FILE)


def get_excel_bytes() -> bytes:
    """Return the Excel file as bytes for download."""
    with open(EXCEL_FILE, "rb") as f:
        return f.read()


# ─── Session State Init ───────────────────────────────────────────────────────
if "refresh" not in st.session_state:
    st.session_state.refresh = 0


def trigger_refresh():
    st.session_state.refresh += 1


# ─── Sidebar Navigation ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding: 16px 0 24px 0;">
        <div style="font-size:24px;font-weight:800;letter-spacing:-0.5px;">🐛 BugTracker</div>
        <div style="font-size:11px;color:#6b7280;font-family:'JetBrains Mono',monospace;margin-top:4px;">v1.0 · openpyxl engine</div>
    </div>
    """, unsafe_allow_html=True)

    nav = st.radio(
        "Navigation",
        ["📊 Dashboard", "➕ Add Bug", "✏️ Edit Bug", "🗑️ Delete Bug"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown("""
    <div style="font-size:11px;color:#6b7280;font-family:'JetBrains Mono',monospace;line-height:1.8;">
    FILE<br>
    <span style="color:#e8eaf0;">""" + EXCEL_FILE + """</span><br><br>
    SHEET<br>
    <span style="color:#e8eaf0;">""" + SHEET_NAME + """</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    with open(EXCEL_FILE, "rb") if os.path.exists(EXCEL_FILE) else open(EXCEL_FILE, "wb") as f:
        pass
    ensure_excel_exists()
    st.download_button(
        "⬇️ Download Excel",
        data=get_excel_bytes(),
        file_name="Book1.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


# ─── Load Data ────────────────────────────────────────────────────────────────
df = read_data()

# ─── Computed Metrics ─────────────────────────────────────────────────────────
total = len(df)
completed = df["Date of complete"].notna().sum() if total > 0 else 0
pending = total - completed
cat_counts = df["Category"].value_counts().to_dict() if total > 0 else {}


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if nav == "📊 Dashboard":
    st.markdown("""
    <div class="app-title">Bug <span class="accent-text">Tracker</span></div>
    <div class="app-subtitle">Real-time · Excel-backed · openpyxl engine</div>
    """, unsafe_allow_html=True)

    # KPI Cards
    cat_summary_text = " · ".join([f"{k}: {v}" for k, v in cat_counts.items()]) or "—"
    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card total">
            <div class="kpi-label">Total Bugs</div>
            <div class="kpi-value">{total}</div>
            <div class="kpi-sub">all records</div>
        </div>
        <div class="kpi-card completed">
            <div class="kpi-label">Completed</div>
            <div class="kpi-value">{completed}</div>
            <div class="kpi-sub">{round(completed/total*100) if total else 0}% done</div>
        </div>
        <div class="kpi-card pending">
            <div class="kpi-label">Pending</div>
            <div class="kpi-value">{pending}</div>
            <div class="kpi-sub">no completion date</div>
        </div>
        <div class="kpi-card category">
            <div class="kpi-label">Categories</div>
            <div class="kpi-value">{len(cat_counts)}</div>
            <div class="kpi-sub">{cat_summary_text}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Filters
    st.markdown('<div class="section-header"><span class="dot"></span> Records</div>', unsafe_allow_html=True)
    with st.container():
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            f_bugid = st.text_input("🔍 BugID", placeholder="e.g. 101")
        with fc2:
            f_app = st.text_input("🔍 AppName", placeholder="filter by app")
        with fc3:
            f_eng = st.text_input("🔍 Engineer", placeholder="filter by name")
        with fc4:
            f_cat = st.selectbox("🔍 Category", ["All"] + [c for c in CATEGORY_OPTIONS if c])

    filtered = df.copy()
    if f_bugid.strip():
        try:
            filtered = filtered[filtered["BugID"] == int(f_bugid.strip())]
        except ValueError:
            pass
    if f_app.strip():
        filtered = filtered[filtered["AppName"].astype(str).str.contains(f_app.strip(), case=False, na=False)]
    if f_eng.strip():
        filtered = filtered[filtered["Enigineer"].astype(str).str.contains(f_eng.strip(), case=False, na=False)]
    if f_cat != "All":
        filtered = filtered[filtered["Category"] == f_cat]

    # Style: highlight pending rows
    def style_rows(row):
        is_pending = pd.isna(row["Date of complete"]) or str(row["Date of complete"]).strip() in ("", "NaT", "None", "nan")
        if is_pending:
            return ["background-color: rgba(255,71,87,0.08); color: #ffd0d4"] * len(row)
        else:
            return ["background-color: rgba(0,200,150,0.05); color: #e8eaf0"] * len(row)

    if not filtered.empty:
        styled = filtered.style.apply(style_rows, axis=1)
        st.dataframe(styled, use_container_width=True, height=420)
        st.markdown(f"""
        <div style="font-size:12px;color:#6b7280;font-family:'JetBrains Mono',monospace;margin-top:8px;">
        Showing {len(filtered)} of {total} records &nbsp;·&nbsp;
        <span style="color:#ff4757;">■</span> Pending &nbsp;
        <span style="color:#00c896;">■</span> Completed
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("No records found matching your filters.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ADD BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "➕ Add Bug":
    st.markdown("""
    <div class="app-title">Add <span class="accent-text">New Bug</span></div>
    <div class="app-subtitle">Fill in the form to append a record to the Excel file</div>
    """, unsafe_allow_html=True)

    with st.form("add_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            #bug_id    = st.number_input("BugID", min_value=1, step=1, value=None, placeholder="e.g. 101")
            bug_id = st.text_input("🔍 BugID", placeholder="e.g. 101")
            title     = st.text_input("Title")
            app_name  = st.text_input("AppName")
            engineer  = st.text_input("Enigineer")
            category  = st.selectbox("Category", CATEGORY_OPTIONS)
            date_assign = st.date_input("Date of assign", value=date.today())
        with c2:
            val_result = st.text_input("Validation Result")
            description = st.text_area("Description", height=100)
            challenges  = st.text_area("Challenges", height=68)
            remarks     = st.text_input("Remarks")
            used_cate   = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS)
            date_complete = st.date_input("Date of complete (optional)", value=None)

        empty_field = st.text_input("Empty (optional)")

        submitted = st.form_submit_button("➕ Add Record", use_container_width=True)

    if submitted:
        if not bug_id:
            st.toast("⚠️ BugID is required.", icon="⚠️")
        elif bug_id in df["BugID"].values:
            st.toast(f"⚠️ BugID {bug_id} already exists.", icon="⚠️")
        else:
            row = {
                "BugID": int(bug_id),
                "Title": title,
                "Description": description,
                "AppName": app_name,
                "Validation Result": val_result,
                "Enigineer": engineer,
                "Date of assign": date_assign.strftime("%Y-%m-%d") if date_assign else None,
                "Date of complete": date_complete.strftime("%Y-%m-%d") if date_complete else None,
                "Empty": empty_field or None,
                "Category": category or None,
                "Challenges": challenges,
                "Remarks": remarks,
                "Used Cate(Yes/No)": used_cate or None,
            }
            try:
                append_row(row)
                st.toast(f"✅ Bug #{int(bug_id)} added successfully!", icon="✅")
                trigger_refresh()
                st.rerun()
            except Exception as e:
                st.toast(f"❌ Error: {e}", icon="❌")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EDIT BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "✏️ Edit Bug":
    st.markdown("""
    <div class="app-title">Edit <span class="accent-text">Bug</span></div>
    <div class="app-subtitle">Select a record by BugID and update its values in-place</div>
    """, unsafe_allow_html=True)

    if df.empty:
        st.info("No records available to edit.")
    else:
        bug_ids = sorted(df["BugID"].dropna().astype(int).tolist())
        selected_id = st.selectbox("Select BugID", bug_ids)

        row = df[df["BugID"] == selected_id].iloc[0]

        def safe_date(val):
            if pd.isna(val) or val is None or str(val).strip() in ("", "NaT", "None", "nan"):
                return None
            try:
                return pd.to_datetime(val).date()
            except Exception:
                return None

        with st.form("edit_form"):
            c1, c2 = st.columns(2)
            with c1:
                title      = st.text_input("Title", value=str(row["Title"]) if pd.notna(row["Title"]) else "")
                app_name   = st.text_input("AppName", value=str(row["AppName"]) if pd.notna(row["AppName"]) else "")
                engineer   = st.text_input("Enigineer", value=str(row["Enigineer"]) if pd.notna(row["Enigineer"]) else "")
                cat_idx    = CATEGORY_OPTIONS.index(row["Category"]) if row["Category"] in CATEGORY_OPTIONS else 0
                category   = st.selectbox("Category", CATEGORY_OPTIONS, index=cat_idx)
                date_assign = st.date_input("Date of assign", value=safe_date(row["Date of assign"]) or date.today())
            with c2:
                val_result  = st.text_input("Validation Result", value=str(row["Validation Result"]) if pd.notna(row["Validation Result"]) else "")
                description = st.text_area("Description", value=str(row["Description"]) if pd.notna(row["Description"]) else "", height=100)
                challenges  = st.text_area("Challenges", value=str(row["Challenges"]) if pd.notna(row["Challenges"]) else "", height=68)
                remarks     = st.text_input("Remarks", value=str(row["Remarks"]) if pd.notna(row["Remarks"]) else "")
                uc_idx      = USED_CATE_OPTIONS.index(row["Used Cate(Yes/No)"]) if row["Used Cate(Yes/No)"] in USED_CATE_OPTIONS else 0
                used_cate   = st.selectbox("Used Cate (Yes/No)", USED_CATE_OPTIONS, index=uc_idx)
                date_complete = st.date_input("Date of complete (optional)", value=safe_date(row["Date of complete"]))

            empty_field = st.text_input("Empty", value=str(row["Empty"]) if pd.notna(row["Empty"]) else "")

            save = st.form_submit_button("💾 Save Changes", use_container_width=True)

        if save:
            updated = {
                "Title": title,
                "Description": description,
                "AppName": app_name,
                "Validation Result": val_result,
                "Enigineer": engineer,
                "Date of assign": date_assign.strftime("%Y-%m-%d") if date_assign else None,
                "Date of complete": date_complete.strftime("%Y-%m-%d") if date_complete else None,
                "Empty": empty_field or None,
                "Category": category or None,
                "Challenges": challenges,
                "Remarks": remarks,
                "Used Cate(Yes/No)": used_cate or None,
            }
            try:
                update_row(selected_id, updated)
                st.toast(f"✅ Bug #{selected_id} updated!", icon="✅")
                trigger_refresh()
                st.rerun()
            except Exception as e:
                st.toast(f"❌ Error: {e}", icon="❌")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DELETE BUG
# ══════════════════════════════════════════════════════════════════════════════
elif nav == "🗑️ Delete Bug":
    st.markdown("""
    <div class="app-title">Delete <span class="accent-text">Bug</span></div>
    <div class="app-subtitle">Permanently remove a record from the Excel file</div>
    """, unsafe_allow_html=True)

    if df.empty:
        st.info("No records available to delete.")
    else:
        bug_ids = sorted(df["BugID"].dropna().astype(int).tolist())
        del_id = st.selectbox("Select BugID to delete", bug_ids)

        # Preview the record
        rec = df[df["BugID"] == del_id]
        st.markdown("**Record preview:**")
        st.dataframe(rec, use_container_width=True)

        confirm = st.checkbox(f"✅ I confirm I want to permanently delete Bug #{del_id}")

        if st.button("🗑️ Delete Record", disabled=not confirm):
            try:
                delete_row(del_id)
                st.toast(f"🗑️ Bug #{del_id} deleted.", icon="🗑️")
                trigger_refresh()
                st.rerun()
            except Exception as e:
                st.toast(f"❌ Error: {e}", icon="❌")
