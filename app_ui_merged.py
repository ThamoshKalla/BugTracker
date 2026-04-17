
# (Merged UI version with requested updates)
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import date
import os

EXCEL_FILE = "Book1.xlsx"
SHEET_NAME = "Sheet2"

COLUMNS = [
    "BugID","Title","Description","AppName",
    "Validation Result","OS Build","Current Status",
    "Enigineer","Date of assign","Date of complete","Empty",
    "Category","Cate Challenges","Challenges","Remarks","Used Cate(Yes/No)"
]

CATEGORY_OPTIONS = ["Anomaly Detection","aetriage","Compatemerging","OCVRegularSerach","Other"]
USED_CATE_OPTIONS = ["","Yes","No"]

VALIDATION_OPTIONS = ["Repro","Not Repro"]
STATUS_OPTIONS = ["Completed","On Hold","Pending","Yet to Start","In Progress"]
ENGINEERS = ["Thomas","Anshika","John"]

st.set_page_config(page_title="Bug Tracker",layout="wide")

def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)

def read_data():
    ensure_excel_exists()
    df = pd.read_excel(EXCEL_FILE)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNS]

def append_row(row):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([row.get(col) for col in COLUMNS])
    wb.save(EXCEL_FILE)

def update_row(bug_id,data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    header=[c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2):
        if str(r[0].value)==str(bug_id):
            for k,v in data.items():
                if k in header:
                    r[header.index(k)].value=v
    wb.save(EXCEL_FILE)

df = read_data()

nav = st.sidebar.radio("Navigation",["📊 Dashboard","➕ Add Bug","✏️ Edit Bug"])

if nav=="➕ Add Bug":
    with st.form("add"):
        c1,c2 = st.columns(2)

        with c1:
            bug_id = st.text_input("BugID")
            title = st.text_input("Title")
            app = st.text_input("AppName")
            engineer = st.selectbox("Engineer",ENGINEERS)
            category = st.selectbox("Category",CATEGORY_OPTIONS)

        with c2:
            val = st.selectbox("Validation Result",VALIDATION_OPTIONS)
            osb = st.text_input("OS Build")
            status = st.selectbox("Current Status",STATUS_OPTIONS)
            used = st.selectbox("Used Cate",USED_CATE_OPTIONS)

        cate_chal=""
        if used=="No":
            cate_chal = st.text_area("Cate Challenges")

        submit = st.form_submit_button("Add")

        if submit:
            append_row({
                "BugID":int(bug_id),
                "Title":title,
                "AppName":app,
                "Validation Result":val,
                "OS Build":osb,
                "Current Status":status,
                "Enigineer":engineer,
                "Date of assign":None,
                "Category":category,
                "Cate Challenges":cate_chal,
                "Used Cate(Yes/No)":used
            })
            st.success("Added")

elif nav=="✏️ Edit Bug":
    ids = df["BugID"].dropna().tolist()
    sel = st.selectbox("BugID",ids)
    row = df[df["BugID"]==sel].iloc[0]

    with st.form("edit"):
        val = st.selectbox("Validation Result",VALIDATION_OPTIONS)
        osb = st.text_input("OS Build",row.get("OS Build",""))
        status = st.selectbox("Current Status",STATUS_OPTIONS)
        eng = st.selectbox("Engineer",ENGINEERS)
        used = st.selectbox("Used Cate",USED_CATE_OPTIONS)

        cate_chal=""
        if used=="No":
            cate_chal = st.text_area("Cate Challenges",row.get("Cate Challenges",""))

        d_assign = st.date_input("Date of assign")

        save = st.form_submit_button("Save")

        if save:
            update_row(sel,{
                "Validation Result":val,
                "OS Build":osb,
                "Current Status":status,
                "Enigineer":eng,
                "Date of assign":d_assign.strftime("%Y-%m-%d"),
                "Cate Challenges":cate_chal,
                "Used Cate(Yes/No)":used
            })
            st.success("Updated")

else:
    st.dataframe(df)
