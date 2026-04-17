
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import date
import os

EXCEL_FILE = "Book1.xlsx"
SHEET_NAME = "Sheet2"

COLUMNS = [
    "BugID", "Title", "Description", "AppName",
    "Validation Result", "OS Build", "Current Status",
    "Enigineer", "Date of assign", "Date of complete",
    "Category", "Cate Challenges", "Challenges",
    "Remarks", "Used Cate(Yes/No)"
]

VALIDATION_OPTIONS = ["Repro", "Not Repro"]
STATUS_OPTIONS = ["Completed", "On Hold", "Pending", "Yet to Start", "In Progress"]
ENGINEERS = ["Thomas", "Anshika", "John", "David"]
USED_CATE_OPTIONS = ["Yes", "No"]
CATEGORY_OPTIONS = ["Anomaly Detection", "aetriage", "Compatemerging", "OCVRegularSerach","Other"]

st.set_page_config(page_title="Bug Tracker", layout="wide")

def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)

def read_data():
    ensure_excel_exists()
    return pd.read_excel(EXCEL_FILE)

def append_row(row):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([row.get(col) for col in COLUMNS])
    wb.save(EXCEL_FILE)

def update_row(bug_id, data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2):
        if str(r[0].value) == str(bug_id):
            for k,v in data.items():
                if k in header:
                    r[header.index(k)].value = v
    wb.save(EXCEL_FILE)

df = read_data()

page = st.sidebar.radio("Menu", ["Add Bug","Edit Bug","View"])

if page == "Add Bug":
    st.header("Add Bug")

    bug_id = st.text_input("BugID")
    title = st.text_input("Title")
    app = st.text_input("AppName")

    val_result = st.selectbox("Validation Result", VALIDATION_OPTIONS)
    os_build = st.text_input("OS Build")
    status = st.selectbox("Current Status", STATUS_OPTIONS)
    engineer = st.selectbox("Engineer", ENGINEERS)

    category = st.selectbox("Category", CATEGORY_OPTIONS)

    used_cate = st.selectbox("Used Cate", USED_CATE_OPTIONS)
    cate_challenges = ""
    if used_cate == "No":
        cate_challenges = st.text_area("Cate Challenges")

    challenges = st.text_area("Challenges")
    remarks = st.text_input("Remarks")

    if st.button("Add"):
        row = {
            "BugID": int(bug_id),
            "Title": title,
            "Description": "",
            "AppName": app,
            "Validation Result": val_result,
            "OS Build": os_build,
            "Current Status": status,
            "Enigineer": engineer,
            "Date of assign": None,
            "Date of complete": None,
            "Category": category,
            "Cate Challenges": cate_challenges if used_cate=="No" else None,
            "Challenges": challenges,
            "Remarks": remarks,
            "Used Cate(Yes/No)": used_cate
        }
        append_row(row)
        st.success("Added")

elif page == "Edit Bug":
    st.header("Edit Bug")

    bug_ids = df["BugID"].dropna().tolist()
    selected = st.selectbox("BugID", bug_ids)

    row = df[df["BugID"]==selected].iloc[0]

    val_result = st.selectbox("Validation Result", VALIDATION_OPTIONS)
    os_build = st.text_input("OS Build", row.get("OS Build",""))
    status = st.selectbox("Current Status", STATUS_OPTIONS)
    engineer = st.selectbox("Engineer", ENGINEERS)

    used_cate = st.selectbox("Used Cate", USED_CATE_OPTIONS)
    cate_challenges = ""
    if used_cate == "No":
        cate_challenges = st.text_area("Cate Challenges")

    date_assign = st.date_input("Date of assign")

    if st.button("Update"):
        update_row(selected,{
            "Validation Result": val_result,
            "OS Build": os_build,
            "Current Status": status,
            "Enigineer": engineer,
            "Date of assign": date_assign.strftime("%Y-%m-%d"),
            "Cate Challenges": cate_challenges,
            "Used Cate(Yes/No)": used_cate
        })
        st.success("Updated")

else:
    st.dataframe(df)
